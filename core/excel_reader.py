"""
Excel文件读取模块 - Flet优化版
"""

import os
import pandas as pd
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook

from utils.cleaner import clean_field_name


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.filename = os.path.basename(file_path)
        self._sheets_info = None
        self._headers_cache = {}
        self._data_cache = {}
    
    def get_sheets(self) -> List[Dict[str, Any]]:
        """获取所有Sheet信息"""
        if self._sheets_info is not None:
            return self._sheets_info
        
        self._sheets_info = []
        
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = ws.max_row if ws.max_row else 0
                try:
                    has_merged = len(ws.merged_cells.ranges) > 0 if hasattr(ws, 'merged_cells') and ws.merged_cells else False
                except Exception:
                    has_merged = False
                self._sheets_info.append({
                    'name': sheet_name,
                    'row_count': row_count,
                    'has_merged_cells': has_merged
                })
            wb.close()
        except Exception as e:
            print(f"Error reading sheets: {e}")
            raise
        
        return self._sheets_info
    
    def get_headers(self, sheet_name: str, header_rows: int = 1) -> List[str]:
        """获取表头"""
        cache_key = f"{sheet_name}_{header_rows}"
        if cache_key in self._headers_cache:
            return self._headers_cache[cache_key]
        
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=list(range(header_rows)),
                nrows=0,
                engine='openpyxl'
            )
            
            headers = []
            if isinstance(df.columns, pd.MultiIndex):
                for col in df.columns:
                    col_name = col[0] if col[0] else ''
                    if not col_name:
                        for c in col:
                            if c:
                                col_name = c
                                break
                    headers.append(clean_field_name(col_name))
            else:
                headers = [clean_field_name(h) for h in df.columns.tolist()]
            
            headers = self._handle_duplicate_headers(headers)
            self._headers_cache[cache_key] = headers
            return headers
            
        except Exception as e:
            print(f"Error reading headers: {e}")
            raise
    
    def _handle_duplicate_headers(self, headers: List[str]) -> List[str]:
        """处理重复的列名"""
        seen = {}
        result = []
        
        for h in headers:
            if h in seen:
                seen[h] += 1
                result.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                result.append(h)
        
        return result
    
    def read_sheet(self, sheet_name: str, header_rows: int = 1, 
                   usecols: list = None) -> pd.DataFrame:
        """读取Sheet数据"""
        cache_key = f"{sheet_name}_{header_rows}"
        if cache_key in self._data_cache:
            return self._data_cache[cache_key]
        
        try:
            read_kwargs = {
                'sheet_name': sheet_name,
                'header': list(range(header_rows)),
                'engine': 'openpyxl',
            }
            if usecols:
                read_kwargs['usecols'] = usecols
            
            df = pd.read_excel(self.file_path, **read_kwargs)
            df.columns = [clean_field_name(col) for col in df.columns]
            
            # 限制缓存大小
            if len(self._data_cache) >= 3:
                oldest_key = next(iter(self._data_cache))
                del self._data_cache[oldest_key]
            
            self._data_cache[cache_key] = df
            return df
            
        except Exception as e:
            print(f"Error reading sheet data: {e}")
            raise
    
    def get_field_info(self, sheet_name: str, max_rows: int = 5) -> List[Dict[str, str]]:
        """获取字段详细信息"""
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                nrows=max_rows,
                engine='openpyxl'
            )
            df.columns = [clean_field_name(col) for col in df.columns]
        except Exception:
            df = pd.DataFrame()
        
        fields = []
        for col in df.columns:
            fields.append({
                'name': col,
                'type': str(df[col].dtype),
                'nullable': df[col].isna().any() if len(df) > 0 else True
            })
        
        return fields
    
    def preview_data(self, sheet_name: str, nrows: int = 10) -> pd.DataFrame:
        """预览数据"""
        df = pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            nrows=nrows,
            engine='openpyxl'
        )
        df.columns = [clean_field_name(col) for col in df.columns]
        return df
    
    def get_summary(self) -> Dict[str, Any]:
        """获取文件摘要信息"""
        size = os.path.getsize(self.file_path)
        
        return {
            'filename': self.filename,
            'path': self.file_path,
            'size': size,
            'size_mb': round(size / (1024 * 1024), 2),
            'sheet_count': len(self.get_sheets()),
            'sheets': self.get_sheets()
        }


def create_reader(file_path: str) -> ExcelReader:
    """工厂函数 - 创建ExcelReader实例"""
    return ExcelReader(file_path)